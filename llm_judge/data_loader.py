"""
Load the collated deduped xlsx into a flat list of evaluation items.

Each item pairs a question row with its answer rows so the LLM judge
sees the same context a human annotator would.
"""

from dataclasses import dataclass, field
from typing import Optional


@dataclass
class EvalItem:
    """One row that needs judging (either a question or an answer)."""
    identifier: str
    language: str
    topic_key: str
    original_text: str
    machine_translation: str
    row_type: str  # "question" or "answer"
    # Context: for answers, store the parent question
    parent_question_original: Optional[str] = None
    parent_question_translation: Optional[str] = None
    # Human labels (filled after humans annotate; None until then)
    human_edit_required: Optional[str] = None
    human_post_edit: Optional[str] = None
    human_cmqm_class: Optional[str] = None
    human_notes: Optional[str] = None


def load_xlsx(xlsx_path: str, languages: list[str] | None = None,
              limit: int | None = None) -> list[EvalItem]:
    """
    Load eval items from the collated deduped workbook.

    Args:
        xlsx_path: Path to llama_translation_collated_deduped.xlsx
        languages: List of language sheet names to load (None = all)
        limit: Max items per language (None = all)

    Returns:
        Flat list of EvalItem objects.
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "The 'openpyxl' package is required to load Excel workbooks. "
            "Install it in the active Python environment."
        ) from exc

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    sheet_names = languages or wb.sheetnames
    items = []

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            print(f"WARNING: sheet '{sheet_name}' not found, skipping")
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        # Track current question for context
        current_q_original = None
        current_q_translation = None
        count = 0

        for row in rows:
            if limit and count >= limit:
                break

            identifier = str(row[0]) if row[0] else ""
            original = str(row[1]) if row[1] else ""
            translation = str(row[2]) if row[2] else ""
            language = str(row[3]) if row[3] else sheet_name
            topic_key = str(row[4]) if row[4] else ""

            # Human annotations (may be None if not yet done)
            edit_req = str(row[5]) if row[5] else None
            post_edit = str(row[6]) if row[6] else None
            cmqm_class = str(row[7]) if row[7] else None
            notes = str(row[8]) if row[8] else None

            # Determine row type from identifier pattern
            # Identifiers like FQ1, SQ1 = questions; FA1.1, SA1.1 = answers
            is_question = "Q" in identifier and "." not in identifier
            row_type = "question" if is_question else "answer"

            if is_question:
                current_q_original = original
                current_q_translation = translation

            item = EvalItem(
                identifier=identifier,
                language=language,
                topic_key=topic_key,
                original_text=original,
                machine_translation=translation,
                row_type=row_type,
                parent_question_original=None if is_question else current_q_original,
                parent_question_translation=None if is_question else current_q_translation,
                human_edit_required=edit_req,
                human_post_edit=post_edit,
                human_cmqm_class=cmqm_class,
                human_notes=notes,
            )
            items.append(item)
            count += 1

    wb.close()
    return items


def load_items_grouped(xlsx_path: str, languages: list[str] | None = None,
                       limit: int | None = None) -> dict[str, list[EvalItem]]:
    """Load items grouped by language."""
    all_items = load_xlsx(xlsx_path, languages, limit)
    grouped: dict[str, list[EvalItem]] = {}
    for item in all_items:
        grouped.setdefault(item.language, []).append(item)
    return grouped
